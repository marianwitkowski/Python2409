
# OpenPyXL
import openpyxl
from openpyxl.styles import Font

# READ XLS
wb = openpyxl.load_workbook("../misc/uod-kopia.xlsx")
sheet = wb.active
print(f"Wierszy={sheet.max_row}, Kolumns={sheet.max_column}")

cell = sheet.cell(row=2, column=1)
print(sheet.cell(row=2, column=1).value) # 191/2021
print(sheet["A2"].value)

# WRITE XLS
wb = openpyxl.Workbook()
sheet = wb.active
for x in range(1,11):
    sheet[f"C{x}"].value = x

# zmiana szerokosci kolumny
sheet.column_dimensions["C"].width = 250

# stylowanie
sheet["C5"].font = Font(size=26, bold=True, name="Calibri")

# nowy arkusz
wb.create_sheet("Nowy testowy")

wb.save("test.xlsx")

